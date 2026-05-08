import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Screening Summary"

# Define colors
green_fill = PatternFill(start_color="2F7F3F", end_color="2F7F3F", fill_type="solid")
blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
red_fill = PatternFill(start_color="C5504C", end_color="C5504C", fill_type="solid")
gray_fill = PatternFill(start_color="505050", end_color="505050", fill_type="solid")
light_blue_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")

white_font = Font(bold=True, color="FFFFFF", size=12)
header_font = Font(bold=True, size=14, color="FFFFFF")
title_font = Font(bold=True, size=16)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws['A1'] = "Screening Report: Fundraising & Partnerships Manager"
ws['A1'].font = title_font
ws.merge_cells('A1:L1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws['A2'] = "Taleemabad Talent Acquisition Agent • 05 March 2026"
ws['A2'].font = Font(italic=True, size=10)
ws.merge_cells('A2:L2')
ws['A2'].alignment = Alignment(horizontal='center')

# Summary stats
ws['A4'] = "Total Applications:"
ws['B4'] = 64
ws['C4'] = "Assessed:"
ws['D4'] = 48
ws['E4'] = "Shortlisted:"
ws['F4'] = 5
ws['G4'] = "Extended Review:"
ws['H4'] = 5
ws['I4'] = "No-Hire:"
ws['J4'] = 38
ws['K4'] = "Out of Budget:"
ws['L4'] = 6

for col in ['A', 'C', 'E', 'G', 'I', 'K']:
    ws[f'{col}4'].font = Font(bold=True)

ws['A5'] = "Budget:"
ws['B5'] = "PKR 150,000 – 270,000 / month"
ws.merge_cells('B5:D5')

# Data for shortlisted candidates
shortlisted_data = [
    ["Danish Hussain", 97.5, "Tier A", "Out of Budget", "PKR 550,000/month", "~20 years",
     "Head of Grants & Partnerships, INGO (Hyderabad)",
     "PKR 1B+ mobilised — FCDO, World Bank, UNDP, ADB named. Highest single fundraising track record in pool.",
     "RECOMMEND — budget exception required"],

    ["Zain Ul Abideen", 95.0, "Tier A", "Out of Budget", "PKR 350,000/month", "~10 years",
     "Deputy Manager Resource Mobilisation, READ Foundation (Islamabad)",
     "US $50M lifetime across two orgs (READ + SPO). Dual-org sequential success confirms repeatable system. Islamabad-based, smallest budget gap of any OOB candidate.",
     "INTERVIEW — negotiate salary"],

    ["Mizhgan Kirmani", 78.8, "Tier B", "In Budget", "PKR 250,000/month", "~8 years",
     "Manager, Donor Relations — The Citizens Foundation (TCF), Islamabad",
     "In-budget, Islamabad-based, active FCDO/UN Women/UNDP/USAID/Green Climate Fund relationships. PKR 72M closed in FY. Best risk-adjusted hire in pool.",
     "INTERVIEW — priority"],

    ["Arsalan Ashraf", 72.2, "Tier B", "Out of Budget", "PKR 450,000/month", "~12 years",
     "Director of Fundraising & Business Development, NGO (Karachi)",
     "Built fundraising functions from scratch at 3 separate NGOs — clearest builder track record for a zero-to-one mandate.",
     "CONSIDER — confirm relocation + negotiate"],

    ["Sadia Sohail", 57.3, "Tier C", "In Budget", "PKR 140,000/month", "~8 years",
     "Donor Relations Officer, READ Foundation (Islamabad)",
     "Most affordable in-budget option. 8 years donor relations at READ Foundation (Taleemabad peer). 3 fundraising certifications.",
     "LOW PRIORITY — growth hire only"],
]

extended_review_data = [
    ["Arsim Tariq", 49.2, "Extended Review", "Borderline", "PKR 280,000–300,000/month", "~10 years",
     "Programme Manager / M&E; Lead (FCDO & World Bank projects, Islamabad)",
     "Deepest sector context in Extended Review — 10 years on FCDO and World Bank-funded contracts; contributed to winning proposals.",
     "EXTENDED REVIEW"],

    ["Ahmed Al-Mayadeen", 45.5, "Extended Review", "Out of Budget", "~PKR 980,000/month", "~10 years",
     "Fundraising Lead, International NGO (Yemen-based)",
     "Elite international fundraising — UN agencies, multi-million dollar scale. Harvard Business School credential. Strongest functional scores after Danish and Zain.",
     "NOT VIABLE — geo + budget"],

    ["Ahad Ahsan Khan", 41.8, "Extended Review", "Out of Budget", "PKR 550,000/month", "~9 years",
     "Manager Grants, Aga Khan University (AKU), Islamabad",
     "$134M portfolio across 210 active grants at AKU. World Bank HEDP lead. Deepest grants compliance expertise in pool.",
     "NOT SUITABLE — wrong function"],

    ["Muhammad Usman", 36.1, "Extended Review", "Out of Budget", "PKR 350,000/month", "~18 years",
     "Public Affairs & Development Alliances Lead (Rawalpindi/Islamabad)",
     "18 years government relations and international development alliance experience. Islamabad-area based.",
     "NOT SUITABLE — no evidence"],

    ["Mushahid Hussain", 34.4, "Extended Review", "In Budget", "PKR 170,000/month", "~4 years",
     "Donor Reporting Officer, READ Foundation (Islamabad)",
     "READ Foundation pedigree. In-budget at PKR 170K. Manages financial reporting to donors.",
     "NOT SUITABLE — junior/reporting only"],
]

# Add headers for shortlisted section
ws['A7'] = "SHORTLISTED CANDIDATES"
ws['A7'].font = white_font
ws['A7'].fill = green_fill
ws.merge_cells('A7:I7')

headers = ["Candidate", "Score", "Tier", "Budget Status", "Expected Salary", "Experience", "Current Role", "Key Strength", "Verdict"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=8, column=col)
    cell.value = header
    cell.font = white_font
    cell.fill = green_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Add shortlisted data
row = 9
for candidate in shortlisted_data:
    for col, value in enumerate(candidate, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = thin_border
        if col == 2:  # Score column
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    row += 1

# Extended review section
extended_start = row + 2
ws[f'A{extended_start}'] = "EXTENDED REVIEW"
ws[f'A{extended_start}'].font = white_font
ws[f'A{extended_start}'].fill = gray_fill
ws.merge_cells(f'A{extended_start}:I{extended_start}')

headers_row = extended_start + 1
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=headers_row, column=col)
    cell.value = header
    cell.font = white_font
    cell.fill = gray_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Add extended review data
row = headers_row + 1
for candidate in extended_review_data:
    for col, value in enumerate(candidate, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = thin_border
        if col == 2:  # Score column
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    row += 1

# No-hire section
no_hire_start = row + 2
ws[f'A{no_hire_start}'] = "NO-HIRE (Sample of 38)"
ws[f'A{no_hire_start}'].font = white_font
ws[f'A{no_hire_start}'].fill = red_fill
ws.merge_cells(f'A{no_hire_start}:I{no_hire_start}')

no_hire_headers_row = no_hire_start + 1
for col, header in enumerate(["Candidate", "Score", "Category", "Current Role", "Why Not Shortlisted", "Verdict"], 1):
    cell = ws.cell(row=no_hire_headers_row, column=col)
    cell.value = header
    cell.font = white_font
    cell.fill = red_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

no_hire_data = [
    ["Mohammad Aqeel Qureshi", 33, "No-Hire", "Manager Fundraising, Shifa Foundation",
     "Fundraising in healthcare/WASH domestic NGO — no multilateral/bilateral grant track record. Shifa donor base is domestic/charity, not USAID/FCDO/WB.", "Not Suitable"],
    ["Faheem Baig", 31, "No-Hire", "Programme Implementation Officer, Development Sector NGO",
     "Programme implementation — not fundraising acquisition. No BD ownership evidenced. Delivery is the opposite function from donor acquisition.", "Not Suitable"],
    ["Shahzad Saleem Abbasi", 30, "No-Hire", "Head of BD, Fundraising & Policy — Junior Jinnah Trust",
     "PKR 400M annually — but entirely domestic charity/CSR model. Donors are individuals and corporates, NOT multilateral/bilateral institutional donors.", "Not Suitable"],
    ["Hamdan Ahmad", 29, "No-Hire", "Programme Manager, World Bank-Funded Programme",
     "Programme management on donor contracts — managing an awarded WB grant is not the same as winning one. No independently won proposals cited.", "Not Suitable"],
    ["Shakir Manzoor", 27, "No-Hire", "Fundraising Support / Grants Assistant, Development NGO",
     "Supporting role — no evidence of independently owning acquisition pipeline or closing grants.", "Not Suitable"],
    ["Sarmad Iqbal", 27, "No-Hire", "Head – Policy Advocacy & Strategic Partnerships, Int'l Development Org",
     "B2G/governance strategy — policy advocacy is distinct from multilateral grant proposal writing and closing.", "Not Suitable"],
    ["Bareera Rauf", 22, "No-Hire", "Junior BD/Fundraising Officer, Development NGO",
     "Less than 3 years experience, no independently closed grants, no multilateral donor relationships.", "Not Suitable"],
    ["Mahnoor Mellu", 18, "No-Hire", "Marketing Executive, Corporate Sector",
     "Corporate marketing background — no development sector, no grant writing, no donor relationship experience. Entirely unrelated function.", "Not Suitable"],
    ["Imran Sarwar", 12, "No-Hire", "Sales Manager, FMCG",
     "FMCG sales — no development sector exposure, no institutional fundraising. Applied to wrong role category.", "Not Suitable"],
]

row = no_hire_headers_row + 1
for candidate in no_hire_data:
    for col, value in enumerate(candidate, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = thin_border
        if col == 2:  # Score column
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    row += 1

# Set column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 25
ws.column_dimensions['H'].width = 35
ws.column_dimensions['I'].width = 25

# Create Dimension Scores sheet
ws_dim = wb.create_sheet("Dimension Scores")

dimension_headers = ["Candidate", "Func", "Outcomes", "Env Fit", "Ownership", "Comms", "Skills", "Growth"]
for col, header in enumerate(dimension_headers, 1):
    cell = ws_dim.cell(row=1, column=col)
    cell.value = header
    cell.font = white_font
    cell.fill = blue_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

dimension_data = [
    ["Danish Hussain", 4, 4, 3, 4, 4, 4, 4],
    ["Zain Ul Abideen", 4, 4, 4, 4, 3, 4, 3],
    ["Mizhgan Kirmani", 3, 3, 4, 3, 3, 3, 3],
    ["Arsalan Ashraf", 4, 3, 3, 4, 3, 3, 4],
    ["Sadia Sohail", 3, 2, 3, 3, 2, 3, 2],
    ["Arsim Tariq", 2, 2, 3, 2, 2, 2, 3],
    ["Ahmed Al-Mayadeen", 4, 4, 1, 4, 3, 3, 3],
    ["Ahad Ahsan Khan", 1, 3, 4, 2, 3, 3, 3],
    ["Muhammad Usman", 2, 1, 3, 2, 3, 1, 3],
    ["Mushahid Hussain", 2, 1, 3, 2, 2, 2, 2],
]

for row_idx, candidate in enumerate(dimension_data, 2):
    for col_idx, value in enumerate(candidate, 1):
        cell = ws_dim.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

        # Color scale for scores
        if col_idx > 1:
            if value == 4:
                cell.fill = PatternFill(start_color="1F7F1F", end_color="1F7F1F", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif value == 3:
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif value == 2:
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif value == 1:
                cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

ws_dim.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_dim.column_dimensions[col].width = 12

# Create Recommendations sheet
ws_rec = wb.create_sheet("Next Steps")

rec_title = ws_rec['A1']
rec_title.value = "Recommended Next Steps"
rec_title.font = title_font

recommendations = [
    ["1", "Confirm Mizhgan Kirmani (#3) for priority interview",
     "In-budget, Islamabad-based, active FCDO/UNDP/USAID relationships. Best risk-adjusted hire."],

    ["2", "Escalate Danish Hussain (#1) and Zain Ul Abideen (#2) to leadership",
     "Both over budget but strongest fundraising track records in pool."],

    ["3", "Explore base + performance structure with Zain Ul Abideen",
     "PKR 80K gap is smallest and most negotiable."],

    ["4", "Confirm relocation commitment from Danish Hussain (Hyderabad)",
     "In writing before scheduling interview."],

    ["5", "Re-advertise if Mizhgan Kirmani does not progress",
     "Sadia Sohail (#5) is an affordable in-budget option but requires 6-month ramp-up."],
]

rec_headers = ["#", "Action", "Rationale"]
for col, header in enumerate(rec_headers, 1):
    cell = ws_rec.cell(row=3, column=col)
    cell.value = header
    cell.font = white_font
    cell.fill = blue_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

for row_idx, rec in enumerate(recommendations, 4):
    for col_idx, value in enumerate(rec, 1):
        cell = ws_rec.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws_rec.column_dimensions['A'].width = 5
ws_rec.column_dimensions['B'].width = 35
ws_rec.column_dimensions['C'].width = 60

# Save workbook
wb.save('c:\\Agent Coco\\reports\\Fundraising_Partnerships_Manager_Screening.xlsx')
print("Excel file created: Fundraising_Partnerships_Manager_Screening.xlsx")
