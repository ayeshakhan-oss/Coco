"""
Module 3: Tier Classification & Excel Generation
- Classify candidates into T1, T2, T3 based on scores
- Generate personalized DM templates
- Create formatted Excel file with tier tabs
"""

import csv
import json
import sys
import os
from datetime import datetime
from typing import List, Dict, Tuple

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../..'))
from utils.audit_log import log_sourcing_action

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)


class TierClassifier:
    def __init__(self, role_config: Dict):
        self.role = role_config.get('title', 'Unknown Role')
        self.config = role_config
        self.tier_thresholds = role_config.get('tier_thresholds', {})

    def classify_candidates(self, enriched_candidates: List[Dict]) -> Dict[str, List[Dict]]:
        """
        Classify candidates into T1, T2, T3 tiers.
        Returns: {
            't1': [candidates],
            't2': [candidates],
            't3': [candidates]
        }
        """
        print(f"\n🏆 Classifying {len(enriched_candidates)} candidates into tiers...")

        tiers = {'t1': [], 't2': [], 't3': []}

        for candidate in enriched_candidates:
            tier = self._determine_tier(candidate)
            candidate['tier'] = tier
            tiers[tier].append(candidate)

        print(f"  T1 (Senior): {len(tiers['t1'])}")
        print(f"  T2 (Mid-level): {len(tiers['t2'])}")
        print(f"  T3 (Junior): {len(tiers['t3'])}")

        # Sort within each tier by confidence score
        for tier_name in tiers:
            tiers[tier_name].sort(
                key=lambda x: float(x.get('confidence_score', 0)),
                reverse=True
            )

        return tiers

    def _determine_tier(self, candidate: Dict) -> str:
        """
        Determine tier based on experience years and GitHub activity score.
        """
        exp_years = float(candidate.get('estimated_years', 3))
        github_score = float(candidate.get('github_activity_score', 0))

        t1_threshold = self.tier_thresholds.get('t1', {})
        t2_threshold = self.tier_thresholds.get('t2', {})

        if (exp_years >= t1_threshold.get('min_years_experience', 7) and
            github_score >= t1_threshold.get('min_github_score', 70)):
            return 't1'
        elif (exp_years >= t2_threshold.get('min_years_experience', 5) and
              github_score >= t2_threshold.get('min_github_score', 50)):
            return 't2'
        else:
            return 't3'

    def generate_dm_template(self, candidate: Dict) -> str:
        """
        Generate a personalized LinkedIn DM template for the candidate.
        """
        name = candidate.get('name', 'there').split()[0]  # First name only
        bio = candidate.get('bio', '')
        company = candidate.get('company', '')
        skills = candidate.get('extracted_skills', '')
        tier = candidate.get('tier', 't3')

        # Personalization based on tier and profile
        if tier == 't1':
            opening = f"Hi {name},\n\nI came across your impressive work in {skills.split(',')[0] if skills else 'tech'} and was genuinely impressed by your GitHub contributions."
        elif tier == 't2':
            opening = f"Hi {name},\n\nYour background in {skills.split(',')[0] if skills else 'product and design'} caught our attention, especially your active contributions to the community."
        else:
            opening = f"Hi {name},\n\nWe saw your interest in {skills.split(',')[0] if skills else 'learning and education'} and thought you'd find our mission compelling."

        dm = f"""{opening}

I'm Ayesha, part of the People & Culture team at Taleemabad — we're building AI-powered tools to improve learning quality for teachers and students across Pakistan.

We're looking for a {self.role} who can shape the product strategy and drive user experience for millions of learners. Given your background in {skills.split(',')[0] if skills else 'product leadership'}, I think you'd find the challenge interesting.

Would you be open to a 20-minute conversation to explore? No pressure at all if the timing isn't right.

Warm regards,
Ayesha Khan
People & Culture | Taleemabad
hiring@taleemabad.com
www.taleemabad.com"""

        return dm


class ExcelGenerator:
    def __init__(self, role_config: Dict):
        self.role = role_config.get('title', 'Unknown Role')
        self.config = role_config

    def generate_excel(self, tiered_candidates: Dict[str, List[Dict]], role_slug: str) -> str:
        """
        Generate Excel file with tier tabs + DM templates.
        Returns: Path to Excel file
        """
        print(f"\n📊 Generating Excel file...")

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define colors for tiers
        tier_colors = {
            't1': 'FFD700',  # Gold
            't2': 'C0C0C0',  # Silver
            't3': 'CD7F32'   # Bronze
        }

        tier_names = {
            't1': 'T1 - Senior',
            't2': 'T2 - Mid-level',
            't3': 'T3 - Junior'
        }

        # Create sheets for each tier
        tier_classifier = TierClassifier(self.config)

        for tier_key in ['t1', 't2', 't3']:
            candidates = tiered_candidates.get(tier_key, [])
            if candidates:
                sheet_name = tier_names[tier_key]
                ws = wb.create_sheet(title=sheet_name)

                # Add header row
                headers = [
                    'Name',
                    'GitHub',
                    'LinkedIn',
                    'Company',
                    'Location',
                    'Skills',
                    'Experience',
                    'GitHub Score',
                    'Confidence',
                    'DM Template'
                ]

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=tier_colors[tier_key], end_color=tier_colors[tier_key], fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Add candidate rows
                for row_num, candidate in enumerate(candidates, 2):
                    dm_template = tier_classifier.generate_dm_template(candidate)

                    row_data = [
                        candidate.get('name', ''),
                        candidate.get('github_url', ''),
                        candidate.get('linkedin_url', ''),
                        candidate.get('company', ''),
                        candidate.get('location', ''),
                        candidate.get('extracted_skills', ''),
                        f"{candidate.get('estimated_years', 3)} years",
                        candidate.get('github_activity_score', 0),
                        candidate.get('confidence_score', 0),
                        dm_template
                    ]

                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = value
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

                # Format columns
                ws.column_dimensions['A'].width = 20  # Name
                ws.column_dimensions['B'].width = 30  # GitHub URL
                ws.column_dimensions['C'].width = 30  # LinkedIn URL
                ws.column_dimensions['D'].width = 20  # Company
                ws.column_dimensions['E'].width = 15  # Location
                ws.column_dimensions['F'].width = 20  # Skills
                ws.column_dimensions['G'].width = 12  # Experience
                ws.column_dimensions['H'].width = 12  # GitHub Score
                ws.column_dimensions['I'].width = 12  # Confidence
                ws.column_dimensions['J'].width = 50  # DM Template

                # Set row height for DM column
                for row in ws.iter_rows(min_row=2, max_row=len(candidates)+1):
                    row[9].cell.alignment = Alignment(vertical='top', wrap_text=True)
                    ws.row_dimensions[row[0].row].height = 150  # Tall rows for DM preview

        # Add "Raw Data" sheet with all candidates
        raw_ws = wb.create_sheet(title='Raw Data', index=0)
        all_candidates = []
        for tier_list in tiered_candidates.values():
            all_candidates.extend(tier_list)

        if all_candidates:
            fieldnames = list(all_candidates[0].keys())
            for col_num, field in enumerate(fieldnames, 1):
                cell = raw_ws.cell(row=1, column=col_num)
                cell.value = field
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

            for row_num, candidate in enumerate(all_candidates, 2):
                for col_num, field in enumerate(fieldnames, 1):
                    cell = raw_ws.cell(row=row_num, column=col_num)
                    cell.value = candidate.get(field, '')
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Auto-adjust columns
            for col_num in range(1, len(fieldnames) + 1):
                raw_ws.column_dimensions[get_column_letter(col_num)].width = 15

        # Save workbook
        output_dir = os.path.join(
            os.path.dirname(__file__),
            f'../../output/sourcing/automation/{role_slug}'
        )
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y-%m-%d')
        excel_path = os.path.join(output_dir, f'{timestamp}-FINAL.xlsx')

        wb.save(excel_path)
        print(f"✅ Excel file saved: {excel_path}")

        return excel_path


if __name__ == '__main__':
    print("This module is meant to be called from the main runner.")
