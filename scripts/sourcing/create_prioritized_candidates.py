#!/usr/bin/env python3
"""
Soul Architect - PRIORITIZED by PRODUCT THINKING + BUILDER + HUMAN DEPTH
Reorganized by tier with scoring against: Product Mindset, Builder Orientation, Human-Centered, Comfort w/ Ambiguity
"""

import os, sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../../")))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Prioritized"

header_fill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
tier1_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
tier2_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
tier3_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
tier4_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ["Tier", "Name", "Role", "Company", "Product Signal", "Builder Signal", "Human-Centered Signal", "LinkedIn"]
ws.append(headers)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# TIER 1: STRONGEST - Clear product + builder + human
tier1 = [
    ("TIER 1", "Faizan Hassan", "Product & AI Strategist", "Independent", "AI chatbot-RAG, product discovery to PMF, strategic thinking", "Consulting model, partners with startups to ship", "Value creation focus, user problem-first", "https://www.linkedin.com/in/faizanhassan/"),
    ("TIER 1", "Ali Akram", "Human-Centered Product Designer", "Design + Research Strategy", "Led human-centered products via research & strategy, AI product design", "Designed & shipped AI product interfaces (Material UI, Tailwind)", "Research-driven, psychology-informed design, behavioral thinking", "https://www.linkedin.com/in/allyakram/"),
    ("TIER 1", "Muhammad Hafih", "Product Manager", "Kollab", "Product manager with AI focus, user research + development", "Building AI-based tutoring product, product ownership", "User research expertise, learning systems, education focus", "https://www.linkedin.com/in/hafihshafiq/"),
    ("TIER 1", "Salahuddin Isa", "Product Manager", "EdTech Strategy & Pedagogy", "Bridges pedagogy + technology as product problem, learning systems", "Shipped EdTech products, strategic implementations", "Deep human + learning psychology, pedagogy expertise", "https://www.linkedin.com/in/salahuddinisa/"),
    ("TIER 1", "Syed Hamza Ali", "CTO & Co-Founder", "Kollab Collections", "Co-founder of conversational product, AI product lead", "Built & shipped conversational product from scratch", "Founder mindset, building for users directly", "https://pk.linkedin.com/in/syed-hamza-ali-63ba31275"),
    ("TIER 1", "Jiya Ali", "Co-founder & ML Engineer", "VentHer", "Founder leading product vision + partnerships, technical depth", "Built & scaling product, founder ownership", "Founder mindset, direct stakeholder relationships", "https://www.linkedin.com/in/jiya-ali-2196b81b0/"),
    ("TIER 1", "Muhammad Irfan", "CEO & Founder", "Xeven Solutions", "AI product company, scaled to 200+ team, product direction", "Built & shipping AI chatbot solutions at scale", "Founder scaling conversational AI products", "https://www.linkedin.com/in/immuhammadirfan/"),
    ("TIER 1", "Asim Ghaffar", "Product / AI Lead", "10Pearls Pakistan", "AI product initiatives, cross-regional product thinking, strategy", "Led AI product implementations, enterprise shipping", "Global perspective, product-led thinking in AI", "https://www.linkedin.com/in/aghaffar/"),
]

# TIER 2: STRONG - Clear product + builder OR human
tier2 = [
    ("TIER 2", "Moiz Alam", "Product Design & Innovation", "Arbisoft Juniper Lab", "Incubation lab leader, product innovation systematizer", "Led Juniper incubation, shipped products for startups", "Design thinking, user-focused innovation", "https://www.linkedin.com/in/moiz994/"),
    ("TIER 2", "Dr. Ayesha Khanna", "CEO & Co-Founder", "Addo.ai", "AI company CEO, product vision + strategy, thought leadership", "Founded & scaling AI solutions firm", "Systems thinking, global perspective, AI future vision", "https://www.linkedin.com/in/ayeshakhanna/"),
    ("TIER 2", "Usman Yameen", "Co-Founder & CEO", "Graphiters", "Design + product agency CEO, award-winning", "Built & scaling design + product agency", "Design-led thinking, client + user focus", "https://www.linkedin.com/in/usman-yameen/"),
    ("TIER 2", "Muhammad Abdullah Qureshi", "Product Manager", "9D Technologies", "9+ yrs product management, AI tools expert, data-driven decisions", "Shipped multiple AI products, agile execution", "Data-driven product thinking", "https://www.linkedin.com/in/muhammad-abdullah-qureshi-897054b9/"),
    ("TIER 2", "Faisal Kamran", "Co-Founder & President", "Addo.ai", "Data science + ML founder, product thinking", "Built data science team, shipping AI solutions", "Founder mindset, data-driven", "https://www.linkedin.com/in/faisalkamran/"),
    ("TIER 2", "Sheikh Izhan Ahmed", "Lead Product Designer", "GetLicenced", "Product + UX designer, EdTech + SaaS experience, top 1% mentor", "Shipped EdTech products, SaaS design", "Education focus, user-centered design, mentoring mindset", "https://www.linkedin.com/in/sheikhizhan/"),
    ("TIER 2", "Atif A.", "Product Manager", "Coder | AI DevEx", "Product manager at AI company, developer experience focus", "Shipping AI/DevX products", "Developer-centric thinking", "https://www.linkedin.com/in/ioatif/"),
    ("TIER 2", "Adeel Pirzada", "Lead Software Architect", "PanaceaLogics", "AI solutions architect, product leadership", "Shipping custom AI solutions", "Product + technical balance", "https://www.linkedin.com/in/adeel-pirzada/"),
    ("TIER 2", "Ziad Aslam", "Senior Product Manager", "Folio3 Software", "Senior PM with end-to-end product ownership", "Shipped multiple products", "Ownership-focused PM", "https://pk.linkedin.com/in/ziadaslam"),
    ("TIER 2", "Muneeb Rashid", "AI/ML Engineer Lead", "Arbisoft", "AI team lead, published research, product + engineering", "Led AI team shipping products, published work", "Research + product bridge", "https://www.linkedin.com/in/muneeb-rashid-2a5b31262/"),
    ("TIER 2", "Hasan Zafar", "Digital Transformation Lead", "AI/Cloud/Analytics", "AI strategist, product-led growth, data-driven", "Led cloud + AI implementations", "Strategic product thinking", "https://www.linkedin.com/in/hasanzafar/"),
]

# TIER 3: SOLID - Good product thinking, some signals
tier3 = [
    ("TIER 3", "Wajeeha Khalid", "Product Manager", "Arbisoft", "PM at embedded product company, team coordination", "Product team experience", "Team-focused product work", "https://www.linkedin.com/in/wajeeha-khalid/"),
    ("TIER 3", "Omar Shah", "CEO & Co-founder", "COLABS", "Startup ecosystem CEO, workspace + product", "Built COLABS, $3M funded", "Ecosystem builder thinking", "https://www.linkedin.com/in/omarshah/"),
    ("TIER 3", "Asma Farooq", "Product Designer", "Design Practice", "Product design leadership, design thinking", "Led design + product work", "Design + product integration", "https://www.linkedin.com/in/asmafarooqonline/"),
    ("TIER 3", "Muhammad Qasim", "Senior Product & UX Designer", "Compass Design Co.", "SaaS + product design, Figma expert", "Shipped SaaS products", "Product design discipline", "https://www.linkedin.com/in/uxkasim/"),
    ("TIER 3", "Syed Waqas Ali Burney", "Product Manager", "Google Research", "LUMS, Google Research PM, climate AI", "Google product shipping, research PM", "Research + product bridge", "https://www.linkedin.com/in/swab/"),
    ("TIER 3", "Abdul Sami", "AI Systems Architecture", "Folio3 Software", "AI systems architect, scalable solutions", "Built AI systems", "Architecture + product thinking", "https://pk.linkedin.com/in/abdulsami"),
    ("TIER 3", "Muhammad Jameel", "AI/Engineer", "Xeven Solutions", "AI engineer at chatbot company", "Shipping AI chatbot solutions", "AI product engineering", "https://www.linkedin.com/in/jameel995/"),
    ("TIER 3", "Usama Altaf", "Product & UX Designer", "Design Firm", "Product + UX integration, design thinking", "Shipped product work", "Product-UX bridge", "https://www.linkedin.com/in/usamaaltaf/"),
    ("TIER 3", "Uswa Zarnab", "Designer", "Wisual Co", "Design agency, product collaboration", "Agency product work", "Design + product", "https://www.linkedin.com/in/uswa-zarnab-6832a6197/"),
    ("TIER 3", "Mohammad Mansoor", "Product Manager", "Toptal", "23+ years product (AI, eCommerce, e-gov), Fortune 100", "Shipped enterprise products", "Seasoned product thinking", "https://www.toptal.com/product-managers/resume/mohammad-mansoor"),
]

# TIER 4: TECHNICAL - Strong engineers, product growth potential
tier4 = [
    ("TIER 4", "Muhammad Ejaz", "Software Engineer", "Arbisoft", "Embedded product team, quality-focused", "Shipped embedded products", "Product team experience", "https://www.linkedin.com/in/muhammad-ejaz-376264b9/"),
    ("TIER 4", "Aimen Khalid", "Engineer", "Arbisoft", "Product development, vision alignment", "Product team experience", "Learning product thinking", "https://www.linkedin.com/in/aimencodechronicles/"),
    ("TIER 4", "Shaheer Alam", "Software Engineer", "Arbisoft", "Emerging product thinking", "Learning product discipline", "Emerging mindset", "https://www.linkedin.com/in/shaheer-alam-51b97a213/"),
    ("TIER 4", "Mushahid Hussain", "Senior Engineer", "10Pearls Pakistan", "Technical depth + client relations", "Shipped enterprise work", "Client-focused engineering", "https://pk.linkedin.com/in/mushahidhussain1"),
    ("TIER 4", "Amna A. Mirza", "Product / Engineering", "10Pearls", "Product + engineering bridge", "Shipped products", "Cross-functional experience", "https://www.linkedin.com/in/amna-a-mirza-/"),
    ("TIER 4", "Zubaira Z.", "Engineer / Product", "10Pearls Pakistan", "Product development, user-centric", "Shipped products", "User-focused engineer", "https://www.linkedin.com/in/zubaira-z/"),
    ("TIER 4", "Safdar Imam", "Associate Director", "10Pearls", "Leadership in top software company", "Shipped enterprise products", "Leadership trajectory", "https://www.linkedin.com/in/safdar-imam-9a309b15/"),
    ("TIER 4", "Mansoor Ali", "Engineer", "10Pearls", "Product development mindset", "Product team experience", "Team experience", "https://www.linkedin.com/in/mansoorharoon/"),
    ("TIER 4", "Muhammad Aamir", "Engineer", "10Pearls", "Product development focus", "Shipped products", "Product engineering", "https://www.linkedin.com/in/muhammad-aamir-650a83b/"),
    ("TIER 4", "Bilal Khan", "Senior Software Engineer", "Confiz Pakistan", "Enterprise, Fortune 100 clients", "Enterprise shipping", "Large-scale systems", "https://www.linkedin.com/in/bilal-khan-784776202/"),
    ("TIER 4", "Muhammad Junaid Pahat", "Machine Learning Engineer", "Confiz", "ML/AI, enterprise context", "Enterprise AI work", "ML + product bridge", "https://www.linkedin.com/in/muhammad-junaid-pahat/"),
    ("TIER 4", "Haider Ali", "Software Engineer", "Confiz Solutions", "Enterprise reliability + UX", "Enterprise shipping", "Quality-focused", "https://www.linkedin.com/in/haider-ali-59597951/"),
    ("TIER 4", "Naveed Shahzad", "Software Engineer", "Confiz Solutions", "Product-focused engineer", "Shipped products", "Product mindset", "https://www.linkedin.com/in/naveed-shahzad-3735b6b/"),
    ("TIER 4", "Shahbaz Mahmood Khan", "Engineer", "Confiz", "Enterprise product balance", "Enterprise shipping", "Technical + product", "https://www.linkedin.com/in/shahbazmahmoodkhan/"),
    ("TIER 4", "Hamza Ehtesham Farooq", "Engineer", "Folio3 Software", "AI/product team", "Shipping AI work", "AI engineering", "https://www.linkedin.com/in/ehteshamxa/"),
    ("TIER 4", "Abdur Raoof", "Engineer", "Folio3 Software", "Product focus", "Product team experience", "Product-minded engineer", "https://www.linkedin.com/in/abdulrauf618/"),
    ("TIER 4", "Muhammad Mujtaba Saeed", "AI Engineer", "Folio3 Software", "AI engineer, product team", "Shipping AI work", "AI product engineering", "https://www.linkedin.com/in/mujtaba-saeed-161019/"),
    ("TIER 4", "Muhammad Bilal", "Engineer", "Folio3 Software", "Product development", "Shipping products", "Product team experience", "https://pk.linkedin.com/in/muhammad-bilal-16749754"),
    ("TIER 4", "Muhammad Usman Bashir", "Engineer", "CyMax Technologies", "AI/ICT solutions", "Enterprise AI work", "Technical foundation", "https://www.linkedin.com/in/muhammad-usman-bashir/"),
    ("TIER 4", "Usman Ishaq", "Chief Revenue & Commercial", "CyMax Technologies", "25+ yrs leadership, P&L experience", "Large-scale execution", "Business + technical", "https://www.linkedin.com/in/ishaqusman/"),
    ("TIER 4", "Noman Butt", "Sales & BD Leader", "CyMax Technologies", "Sales + strategy", "Business development", "Opportunity thinking", "https://www.linkedin.com/in/nomankhalidbutt/"),
    ("TIER 4", "Fahd Khan", "Technology Leader", "CyMax Technologies", "Tech leadership, global perspective", "Large-scale systems", "Technical leadership", "https://www.linkedin.com/in/fahd-khan/"),
]

all_tiers = [tier1, tier2, tier3, tier4]
tier_colors = [tier1_fill, tier2_fill, tier3_fill, tier4_fill]

for tier_idx, tier_data in enumerate(all_tiers):
    for row_data in tier_data:
        ws.append(row_data)
        row_num = ws.max_row

        # Color code by tier
        for cell in ws[row_num]:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.fill = tier_colors[tier_idx]

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 38
ws.column_dimensions['F'].width = 38
ws.column_dimensions['G'].width = 38
ws.column_dimensions['H'].width = 42

ws.row_dimensions[1].height = 30

summary = wb.create_sheet("Strategy")
summary['A1'] = "Soul Architect - Prioritized by Product + Builder + Human"
summary['A1'].font = Font(bold=True, size=14, color="1565c0")

summary['A3'] = "TIER 1 (Start here):"
summary['B3'] = "8 candidates - Product PM + Builder + Human depth signals"
summary['A4'] = "TIER 2 (Strong):"
summary['B4'] = "11 candidates - Clear product + builder OR human signals"
summary['A5'] = "TIER 3 (Solid):"
summary['B5'] = "10 candidates - Good product thinking, some signals"
summary['A6'] = "TIER 4 (Technical):"
summary['B6'] = "24 engineers - Strong execution, product growth potential"

summary['A8'] = "Strategy:"
summary['A9'] = "- Start outreach with Tier 1 (8 people) - highest product + builder + human match"
summary['A10'] = "- Follow with Tier 2 (11 people) - strong secondary tier"
summary['A11'] = "- Tier 3 & 4 as secondary options or growth/mentoring roles"

file_path = "c:\\Agent Coco\\output\\sourcing\\Soul_Architect_PRIORITIZED_ByProductThinking_2026-04-16.xlsx"
os.makedirs(os.path.dirname(file_path), exist_ok=True)
wb.save(file_path)

print(f"[CREATED] Prioritized sheet by Product + Builder + Human")
print(f"Tier 1 (Strongest): 8 candidates")
print(f"Tier 2 (Strong): 11 candidates")
print(f"Tier 3 (Solid): 10 candidates")
print(f"Tier 4 (Technical): 24 candidates")
print(f"File: {file_path}")
