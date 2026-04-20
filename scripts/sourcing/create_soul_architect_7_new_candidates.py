#!/usr/bin/env python3
"""
Soul Architect - 7 New Verified Candidates (User-Researched)
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../../")))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "7 New Candidates"

header_fill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
tier_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ["#", "Tier", "Name", "Current Role", "Company", "Location", "Product Thinking Signals", "LinkedIn Profile"]
ws.append(headers)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

candidates = [
    (1, "NEW", "Mohsin Khan", "Product Automation | UIUX Designer", "Design Practice", "Pakistan", "UI/UX design automation, product-focused design", "https://www.linkedin.com/in/mohsin-khan777/"),
    (2, "NEW", "Zohaib Khan", "UI/UX Designer", "Design Agency", "Pakistan", "UI/UX design, graphic design, product mindset", "https://www.linkedin.com/in/zohaib-khan-183840275/"),
    (3, "NEW", "Ahmed Shahwar", "User Research Associate", "AI Company", "Pakistan", "AI automation, user research, product research", "https://www.linkedin.com/in/ahmedshahwar/"),
    (4, "NEW", "Syed Sarib Sultan", "Product Design Specialist", "Design Studio", "Pakistan", "Product design specialization, design systems", "https://www.linkedin.com/in/syedsaribsultanyac270/"),
    (5, "NEW", "Parivash Mir", "UX Designer", "Design Practice", "Pakistan", "UX design, content design, AI product focus", "https://www.linkedin.com/in/parivashmir/"),
    (6, "NEW", "Shafaq Noor", "Brand & Communications", "Brand Agency", "Pakistan", "Communications focus, brand strategy, product positioning", "https://www.linkedin.com/in/shafaq-noor-95a15618b/"),
    (7, "NEW", "Laraib Piracha", "Product & Growth Analyst", "Analytics/Growth", "Pakistan", "Product growth focus, analytics, product strategy", "https://www.linkedin.com/in/laraib-piracha-31819b42/"),
]

for idx, row_data in enumerate(candidates):
    ws.append(row_data)
    row_num = ws.max_row

    for cell in ws[row_num]:
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = tier_fill

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 26
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 40
ws.column_dimensions['H'].width = 48

ws.row_dimensions[1].height = 35

summary = wb.create_sheet("Summary")
summary['A1'] = "Soul Architect - 7 New Verified Candidates"
summary['A1'].font = Font(bold=True, size=14, color="1565c0")

summary['A3'] = "TOTAL CANDIDATES:"
summary['B3'] = "7 (user-researched)"
summary['A4'] = "ALL TIER:"
summary['B4'] = "NEW - Fresh additions for sourcing slate"
summary['A5'] = "LOCATION:"
summary['B5'] = "Pakistan-based"
summary['A6'] = "DATE ADDED:"
summary['B6'] = datetime.now().strftime('%Y-%m-%d')

file_path = "c:\\Agent Coco\\output\\sourcing\\Soul_Architect_7_New_Candidates_2026-04-18.xlsx"
os.makedirs(os.path.dirname(file_path), exist_ok=True)
wb.save(file_path)

print(f"[CREATED] Soul Architect - 7 New Candidates")
print(f"File: {file_path}")
print(f"Total: 7 new candidates")
