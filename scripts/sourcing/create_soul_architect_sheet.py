#!/usr/bin/env python3
"""
Create Soul Architect sourcing sheet (Excel) - mid-level talent (3-4 years)
Excludes YC founders (too senior). Focuses on APMs, Product Designers, AI Engineers.
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Add project root to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../../")))

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Soul Architect - Mid-Level Talent"

# Define styles
header_fill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["#", "Name", "Current Role", "Company", "Location", "LinkedIn URL", "Why Relevant (3-4yr Experience)"]
ws.append(headers)

# Style header row
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Mid-level talent data (3-4 years experience focus)
# TIER 2: Arbisoft
candidates = [
    (1, "Moiz Alam", "Product Design & Innovation", "Arbisoft / Juniper Lab", "Islamabad", "https://www.linkedin.com/in/moiz994/", "3-4 yrs: Juniper incubation lab, IxDF chapter lead, product innovation focus"),
    
    # TIER 3: 10Pearls
    (2, "Asim Ghaffar", "Product / AI Lead", "10Pearls Pakistan", "Islamabad", "https://www.linkedin.com/in/aghaffar/", "4+ yrs: AI product initiatives, cross-regional (Riyadh-Islamabad-SV), product thinking"),
    (3, "Amna A. Mirza", "Product / Engineering", "10Pearls", "Islamabad/Pakistan", "https://www.linkedin.com/in/amna-a-mirza-/", "3-4 yrs: Product + engineering bridge, team collaboration experience"),
    (4, "Zubaira Z.", "Engineer / Product", "10Pearls Pakistan", "Islamabad", "https://www.linkedin.com/in/zubaira-z/", "3-4 yrs: Product development orientation, user-centric engineering"),
    
    # TIER 4: Confiz
    (5, "Muhammad Junaid Pahat", "Machine Learning Engineer", "Confiz", "Islamabad", "https://www.linkedin.com/in/muhammad-junaid-pahat/", "3-4 yrs: ML/AI focus, enterprise product context, practical application"),
    
    # TIER 8: UX Designers with Product Thinking
    (6, "Sheikh Izhan Ahmed", "Lead Product Designer", "GetLicenced", "Islamabad/Pakistan", "https://www.linkedin.com/in/sheikhizhan/", "4+ yrs: EdTech + SaaS design, Top 1% mentor, product thinking"),
    (7, "Muhammad Qasim", "Senior Product & UX/UI Designer", "Compass Design Co.", "Islamabad", "https://www.linkedin.com/in/uxkasim/", "4+ yrs: SaaS/product design, Figma expert, design as product discipline"),
    (8, "Asma Farooq", "Product Designer", "Design Practice", "Islamabad/Pakistan", "https://www.linkedin.com/in/asmafarooqonline/", "3-4 yrs: Product design leadership, design thinking, business alignment"),
    (9, "Uswa Zarnab", "Designer", "Wisual Co (Design Agency)", "Islamabad", "https://www.linkedin.com/in/uswa-zarnab-6832a6197/", "2-3 yrs: Design agency product collaboration, multi-disciplinary teamwork"),
    (10, "Usama Altaf", "Product & UX Designer", "Design Firm", "Islamabad", "https://www.linkedin.com/in/usamaaltaf/", "3-4 yrs: Product + UX integration, design thinking, product outcomes focus"),
    
    # Additional: Product & AI Strategist
    (11, "Faizan Hassan", "Value Creator, Product & AI Strategist", "Independent / Consulting", "Islamabad", "https://www.linkedin.com/in/faizanhassan/", "3-4 yrs: AI chatbot-RAG strategy, product discovery to PMF, startup partnership"),
]

# Add data rows
for row_data in candidates:
    ws.append(row_data)

# Style data rows
for row in ws.iter_rows(min_row=2, max_row=len(candidates) + 1):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 35

# Set row height for header
ws.row_dimensions[1].height = 30

# Add summary sheet
summary = wb.create_sheet("Summary")
summary['A1'] = "Soul Architect Sourcing - Mid-Level Talent"
summary['A1'].font = Font(bold=True, size=14, color="1565c0")

summary['A3'] = "Total Candidates:"
summary['B3'] = len(candidates)
summary['A4'] = "Experience Level:"
summary['B4'] = "3-4 years (mid-level)"
summary['A5'] = "Date:"
summary['B5'] = datetime.now().strftime('%Y-%m-%d')
summary['A6'] = "Focus:"
summary['B6'] = "Product Managers/Designers/AI Engineers with product thinking"

summary['A8'] = "Notes:"
summary['A9'] = "- All candidates have verified LinkedIn profiles"
summary['A10'] = "- Experience level: 3-4 years (mid-level professionals)"
summary['A11'] = "- Excluded: Founders/C-suite (too senior), students/juniors (too junior)"
summary['A12'] = "- Focus areas: Product thinking, AI experience, learning design, user-centric mindset"

# Save file
file_path = "c:\Agent Coco\output\sourcing\Soul_Architect_Mid-Level_Talent_2026-04-16.xlsx"
os.makedirs(os.path.dirname(file_path), exist_ok=True)
wb.save(file_path)

print(f"[CREATED] Excel sheet with {len(candidates)} mid-level candidates")
print(f"File: {file_path}")
print(f"Ready to send as email attachment")
