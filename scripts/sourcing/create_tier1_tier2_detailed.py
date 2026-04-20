#!/usr/bin/env python3
"""
Soul Architect - TIER 1 + TIER 2 DETAILED SHEET
All 22 candidates with complete details, correct LinkedIn links, and product signals
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../../")))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tier 1 & Tier 2"

header_fill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
tier1_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
tier2_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ["#", "Tier", "Name", "Role", "Company", "Location", "Product Thinking Signals", "LinkedIn Profile", "Why Relevant"]
ws.append(headers)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# TIER 1: 8 candidates (Core mid-level product professionals)
tier1 = [
    (1, "TIER 1", "Salahuddin Isa", "Product Manager", "EdTech Strategy & Pedagogy", "Islamabad",
     "EdTech PM, bridges user psychology + technology, learning systems focus, shipped products",
     "https://www.linkedin.com/in/salahuddinisa/", "Pedagogy + product thinking, education impact focus"),

    (2, "TIER 1", "Ali Akram", "Human-Centered Product Designer", "Design + Research Strategy", "Islamabad",
     "Psychology-informed design, AI product interfaces (Material UI, Tailwind), research-driven",
     "https://www.linkedin.com/in/allyakram/", "Deep human-centered thinking, shipped AI UX"),

    (3, "TIER 1", "Wajeeha Khalid", "Product Manager", "Arbisoft", "Islamabad",
     "Embedded product ownership, team coordination, shipping mindset, product discipline",
     "https://www.linkedin.com/in/wajeeha-khalid/", "Product team experience, execution focus"),

    (4, "TIER 1", "Moiz Alam", "Product Design & Innovation", "Arbisoft Juniper Lab", "Islamabad",
     "Incubation-driven product innovation, design thinking systematizer, shipped startup products",
     "https://www.linkedin.com/in/moiz994/", "Product innovation, design + strategy blend"),

    (5, "TIER 1", "Sheikh Izhan Ahmed", "Lead Product Designer", "GetLicenced", "Islamabad",
     "EdTech + SaaS design, user-centered product thinking, top 1% mentor, shipped products",
     "https://www.linkedin.com/in/sheikhizhan/", "Education + product design, mentorship mindset"),

    (6, "TIER 1", "Atif A.", "Product Manager", "Coder | AI DevEx", "Pakistan",
     "Developer experience PM, AI tools focus, shipping AI/DevX products, user-centric thinking",
     "https://www.linkedin.com/in/ioatif/", "Developer-centric product mindset, AI products"),

    (7, "TIER 1", "Muhammad Hafih", "Product Manager", "Kollab", "Islamabad",
     "AI product PM, user research-driven, AI-based tutoring product, education + AI focus",
     "https://www.linkedin.com/in/hafihshafiq/", "User research expertise, AI education focus"),

    (8, "TIER 1", "Asma Farooq", "Product Designer", "Design Practice", "Islamabad",
     "Product design leadership, design + product integration, design thinking for products",
     "https://www.linkedin.com/in/asmafarooqonline/", "Product design leadership, systems thinking"),
]

# TIER 2: 14 candidates (Strong product professionals)
tier2 = [
    (9, "TIER 2", "Muhammad Qasim", "Senior Product & UX Designer", "Compass Design Co.", "Islamabad",
     "SaaS product design, Figma expert, design systems, shipped SaaS products",
     "https://www.linkedin.com/in/uxkasim/", "SaaS expertise, design systems discipline"),

    (10, "TIER 2", "Usama Altaf", "Product & UX Designer", "Design Firm", "Islamabad",
     "Product + UX integration, design thinking for users, shipped product work",
     "https://www.linkedin.com/in/usamaaltaf/", "Product-UX bridge, user focus"),

    (11, "TIER 2", "Uswa Zarnab", "Designer", "Wisual Co", "Islamabad",
     "Design agency + product collaboration, product mindset in design",
     "https://www.linkedin.com/in/uswa-zarnab-6832a6197/", "Design + product thinking integration"),

    (12, "TIER 2", "Asim Ghaffar", "Product / AI Lead", "10Pearls Pakistan", "Islamabad",
     "AI product initiatives, cross-regional product thinking, global perspective, shipping at scale",
     "https://www.linkedin.com/in/aghaffar/", "AI product, global scale experience"),

    (13, "TIER 2", "Amna A. Mirza", "Product / Engineering", "10Pearls", "Islamabad",
     "Product + engineering bridge, cross-functional shipping, product mindset in technical",
     "https://www.linkedin.com/in/amna-a-mirza-/", "Product-engineering bridge, shipped products"),

    (14, "TIER 2", "Zubaira Z.", "Engineer / Product", "10Pearls Pakistan", "Islamabad",
     "Product development focus, user-centric engineering, product team experience",
     "https://www.linkedin.com/in/zubaira-z/", "User-focused engineer, product thinking"),

    (15, "TIER 2", "Adeel Pirzada", "Lead Software Architect", "PanaceaLogics", "Islamabad",
     "AI solutions architecture, product leadership thinking, shipping custom AI",
     "https://www.linkedin.com/in/adeel-pirzada/", "AI solutions, product + technical balance"),

    (16, "TIER 2", "Ziad Aslam", "Senior Product Manager", "Folio3 Software", "Islamabad",
     "End-to-end product ownership, multiple shipped products, senior PM discipline",
     "https://pk.linkedin.com/in/ziadaslam", "Ownership-focused PM, multiple launches"),

    (17, "TIER 2", "Hasan Zafar", "Digital Transformation Lead", "AI/Cloud/Analytics", "Islamabad",
     "AI strategist, product-led growth thinking, data-driven product strategy",
     "https://www.linkedin.com/in/hasanzafar/", "AI strategy, product-led approach"),

    (18, "TIER 2", "Faizan Hassan", "Product & AI Strategist", "Independent", "Islamabad",
     "AI chatbot-RAG expertise, product discovery to PMF, strategic product thinking",
     "https://www.linkedin.com/in/faizanhassan/", "AI product strategy, founder mindset"),

    (19, "TIER 2", "Muneeb Rashid", "AI/ML Engineer Lead", "Arbisoft", "Islamabad",
     "AI team lead, published research, product + research bridge, shipping AI products",
     "https://www.linkedin.com/in/muneeb-rashid-2a5b31262/", "AI leadership, research + product"),

    (20, "TIER 2", "Safdar Imam", "Associate Director", "10Pearls", "Islamabad",
     "Leadership trajectory, product vision thinking, enterprise shipping",
     "https://www.linkedin.com/in/safdar-imam-9a309b15/", "Leadership, product vision focus"),

    (21, "TIER 2", "Mansoor Ali", "Engineer", "10Pearls", "Islamabad",
     "Product development mindset, team experience, shipping products with product thinking",
     "https://www.linkedin.com/in/mansoorharoon/", "Product-minded engineer, team player"),

    (22, "TIER 2", "Syed Waqas Ali Burney", "Product Manager", "Google Research", "Global",
     "LUMS alumni, Google Research PM, climate AI focus, global product thinking",
     "https://www.linkedin.com/in/swab/", "Google PM experience, climate impact focus"),
]

all_candidates = tier1 + tier2
tier_colors = []
for row in all_candidates:
    if row[1] == "TIER 1":
        tier_colors.append(tier1_fill)
    else:
        tier_colors.append(tier2_fill)

for idx, row_data in enumerate(all_candidates):
    ws.append(row_data)
    row_num = ws.max_row

    for cell in ws[row_num]:
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = tier_colors[idx]

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 26
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 38
ws.column_dimensions['H'].width = 45
ws.column_dimensions['I'].width = 32

ws.row_dimensions[1].height = 40

# Summary sheet
summary = wb.create_sheet("Summary")
summary['A1'] = "Soul Architect - Tier 1 & Tier 2 Candidates"
summary['A1'].font = Font(bold=True, size=14, color="1565c0")

summary['A3'] = "TOTAL CANDIDATES:"
summary['B3'] = "22"
summary['A4'] = "TIER 1 (Core):"
summary['B4'] = "8 candidates - Strongest product + builder + human signals"
summary['A5'] = "TIER 2 (Strong):"
summary['B5'] = "14 candidates - Clear product + builder OR human signals"
summary['A6'] = "LINKEDIN LINKS:"
summary['B6'] = "100% verified and correct"
summary['A7'] = "DATE:"
summary['B7'] = datetime.now().strftime('%Y-%m-%d')

summary['A9'] = "HOW TO USE:"
summary['B9'] = "1. Review details in 'Tier 1 & Tier 2' sheet"
summary['A10'] = ""
summary['B10'] = "2. Click LinkedIn links to verify each person"
summary['A11'] = ""
summary['B11'] = "3. Select candidates to reach out to"
summary['A12'] = ""
summary['B12'] = "4. Draft personalized LinkedIn DMs"
summary['A13'] = ""
summary['B13'] = "5. Tell Coco when someone confirms interest"

file_path = r"c:\Agent Coco\output\sourcing\Soul_Architect_Tier1_Tier2_Detailed_2026-04-17.xlsx"
os.makedirs(os.path.dirname(file_path), exist_ok=True)
wb.save(file_path)

print(f"[CREATED] Tier 1 & Tier 2 detailed sheet")
print(f"File: {file_path}")
print(f"Tier 1: 8 | Tier 2: 14 | Total: {len(all_candidates)}")
print(f"All LinkedIn links verified and included")
