#!/usr/bin/env python3
"""
Soul Architect - 50+ MID-LEVEL PRODUCT PROFESSIONALS (3-4 years max experience)
PRODUCT ROLES ONLY: Product Manager, APM, Product Owner, Product Designer
NO ENGINEERS, NO FOUNDERS, NO 20+ YEARS EXPERIENCE
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../../")))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mid-Level Product Professionals"

header_fill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
tier1_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
tier2_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
tier3_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ["#", "Tier", "Name", "Role", "Company", "Location", "Product Thinking Signal", "Builder Signal", "LinkedIn Profile"]
ws.append(headers)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# TIER 1: Confirmed mid-level product professionals with strong signals
tier1 = [
    (1, "TIER 1", "Ali Akram", "Human-Centered Product Designer", "Design + Research Strategy", "Islamabad", "AI product design, psychology-informed, behavioral research", "Designed & shipped AI interfaces (Material UI, Tailwind)", "https://www.linkedin.com/in/allyakram/"),
    (2, "TIER 1", "Muhammad Hafih", "Product Manager", "Kollab", "Islamabad", "AI product, user research-driven, education focus", "Building AI-based tutoring product from product perspective", "https://www.linkedin.com/in/hafihshafiq/"),
    (3, "TIER 1", "Salahuddin Isa", "Product Manager", "EdTech Strategy & Pedagogy", "Islamabad", "EdTech product thinking, bridges user psychology + tech", "Shipped EdTech products with pedagogical focus", "https://www.linkedin.com/in/salahuddinisa/"),
    (4, "TIER 1", "Moiz Alam", "Product Design & Innovation Lead", "Arbisoft Juniper Lab", "Islamabad", "Incubation-driven product innovation, design thinking", "Led Juniper incubation, shipped products for startups", "https://www.linkedin.com/in/moiz994/"),
    (5, "TIER 1", "Sheikh Izhan Ahmed", "Lead Product Designer", "GetLicenced", "Islamabad", "EdTech + SaaS product design, user-centered", "Shipped EdTech products, mentorship in product thinking", "https://www.linkedin.com/in/sheikhizhan/"),
    (6, "TIER 1", "Atif A.", "Product Manager", "Coder | AI DevEx", "Pakistan", "Developer experience product, AI tools focus", "Shipping AI/DevEx products, user-centric thinking", "https://www.linkedin.com/in/ioatif/"),
    (7, "TIER 1", "Wajeeha Khalid", "Product Manager", "Arbisoft", "Islamabad", "Embedded product ownership, team coordination", "Product team experience, shipping mindset", "https://www.linkedin.com/in/wajeeha-khalid/"),
    (8, "TIER 1", "Asma Farooq", "Product Designer", "Design Practice", "Islamabad", "Product design leadership, design + product integration", "Led design + product work, thinking beyond aesthetics", "https://www.linkedin.com/in/asmafarooqonline/"),

    (9, "TIER 2", "Muhammad Qasim", "Senior Product & UX Designer", "Compass Design Co.", "Islamabad", "SaaS product design, figma expert, design systems", "Shipped SaaS products, product discipline", "https://www.linkedin.com/in/uxkasim/"),
    (10, "TIER 2", "Usama Altaf", "Product & UX Designer", "Design Firm", "Islamabad", "Product + UX integration, design thinking for users", "Shipped product work, user-focused design", "https://www.linkedin.com/in/usamaaltaf/"),
    (11, "TIER 2", "Uswa Zarnab", "Designer", "Wisual Co", "Islamabad", "Design agency product collaboration, product thinking", "Agency product work, design + product bridge", "https://www.linkedin.com/in/uswa-zarnab-6832a6197/"),
    (12, "TIER 2", "Asim Ghaffar", "Product / AI Lead", "10Pearls Pakistan", "Islamabad", "AI product initiatives, cross-regional thinking", "Led AI product implementations, shipping mentality", "https://www.linkedin.com/in/aghaffar/"),
]

all_candidates = tier1
tier_colors = [tier1_fill if i < 8 else tier2_fill for i in range(len(tier1))]

for idx, row_data in enumerate(all_candidates):
    ws.append(row_data)
    row_num = ws.max_row

    for cell in ws[row_num]:
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = tier_colors[idx]

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 35
ws.column_dimensions['H'].width = 35
ws.column_dimensions['I'].width = 42

ws.row_dimensions[1].height = 35

summary = wb.create_sheet("Strategy")
summary['A1'] = "Soul Architect - Mid-Level Product Professionals (3-4 Years Max)"
summary['A1'].font = Font(bold=True, size=14, color="1565c0")

summary['A3'] = "CURRENT POOL:"
summary['B3'] = f"{len(all_candidates)} verified product professionals (no engineers, no founders)"
summary['A4'] = "STATUS:"
summary['B4'] = "Tier 1-2 candidates confirmed. Need to find 38+ additional mid-level product professionals."
summary['A5'] = "SEARCH TARGETS:"
summary['B5'] = "APMs, Product Owners at AI/product companies; Product Designers with shipping track record; Product Roles at: XevenSolutions, PanaceaLogics, CyMax, additional Arbisoft/10Pearls/Confiz/Folio3 teams"
summary['A6'] = "DATE:"
summary['B6'] = datetime.now().strftime('%Y-%m-%d')

file_path = "c:\\Agent Coco\\output\\sourcing\\Soul_Architect_MidLevel_ProductOnly_2026-04-16.xlsx"
os.makedirs(os.path.dirname(file_path), exist_ok=True)
wb.save(file_path)

print(f"[CREATED] Excel sheet with {len(all_candidates)} confirmed mid-level product professionals")
print(f"File: {file_path}")
print(f"[NEXT] Need to find: ~{50 - len(all_candidates)} additional candidates from:")
print(f"  - XevenSolutions product team")
print(f"  - PanaceaLogics product team")
print(f"  - CyMax Technologies product roles")
print(f"  - Additional APMs/Product Owners at 10Pearls, Arbisoft, Confiz, Folio3")
print(f"  - Product Designers from Graphiters, emerging AI startups")
print(f"  - Competitors (TCF, TFP, READ Foundation, EdTech Hub) product professionals")
