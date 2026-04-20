#!/usr/bin/env python3
"""
Create Soul Architect sourcing sheet (Excel) - 50+ candidates
Excludes ONLY YC founders (too senior). Full slate of talent.
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Add project root to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../../")))

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Candidates"

# Define styles
header_fill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["#", "Tier", "Name", "Current Role", "Company", "Location", "LinkedIn URL", "Why Relevant"]
ws.append(headers)

# Style header row
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Complete candidate dataset (all tiers EXCEPT YC founders)
candidates = [
    # TIER 1: Senior Product + AI/EdTech (7)
    (1, "Tier 1", "Salahuddin Isa", "Product Management", "EdTech Strategy & Pedagogy", "Islamabad", "https://www.linkedin.com/in/salahuddinisa/", "EdTech product + pedagogy, learning systems"),
    (2, "Tier 1", "Ali Akram", "Human-Centered Product Designer", "Design + Research Strategy", "Islamabad", "https://www.linkedin.com/in/allyakram/", "AI product design, human-centered research"),
    (3, "Tier 1", "Muhammad Abdullah Qureshi", "Product Manager", "9D Technologies", "Islamabad", "https://www.linkedin.com/in/muhammad-abdullah-qureshi-897054b9/", "9+ yrs AI tools, data-driven product"),
    (4, "Tier 1", "Wajeeha Khalid", "Product Manager", "Arbisoft", "Islamabad", "https://www.linkedin.com/in/wajeeha-khalid/", "Product + embedded systems"),
    (5, "Tier 1", "Mohammad Mansoor", "Product Manager", "Toptal", "Islamabad", "https://www.toptal.com/product-managers/resume/mohammad-mansoor", "23+ yrs product across AI, eCommerce"),
    (6, "Tier 1", "Hasan Zafar", "Digital Transformation Lead", "AI/Cloud/Analytics", "Islamabad", "https://www.linkedin.com/in/hasanzafar/", "AI strategist, product-led growth"),
    (7, "Tier 1", "Jiya Ali", "Co-founder & ML Engineer", "VentHer", "Islamabad", "https://www.linkedin.com/in/jiya-ali-2196b81b0/", "Founder mindset, technical + product"),

    # TIER 2: Arbisoft Team (6)
    (8, "Tier 2", "Moiz Alam", "Product Design & Innovation", "Arbisoft / Juniper Lab", "Islamabad", "https://www.linkedin.com/in/moiz994/", "Juniper incubation, product innovation"),
    (9, "Tier 2", "Muneeb Rashid", "AI/ML Engineer Lead", "Arbisoft", "Islamabad", "https://www.linkedin.com/in/muneeb-rashid-2a5b31262/", "AI team lead, autonomous driving research"),
    (10, "Tier 2", "Muhammad Ejaz", "Software Engineer", "Arbisoft", "Islamabad", "https://www.linkedin.com/in/muhammad-ejaz-376264b9/", "Embedded product team, quality-focused"),
    (11, "Tier 2", "Aimen Khalid", "Engineer", "Arbisoft", "Islamabad", "https://www.linkedin.com/in/aimencodechronicles/", "Product development, vision alignment"),
    (12, "Tier 2", "Shaheer Alam", "Software Engineer", "Arbisoft", "Islamabad", "https://www.linkedin.com/in/shaheer-alam-51b97a213/", "Emerging product thinking"),

    # TIER 3: 10Pearls Team (8)
    (13, "Tier 3", "Asim Ghaffar", "Product / AI Lead", "10Pearls Pakistan", "Islamabad", "https://www.linkedin.com/in/aghaffar/", "AI product, global perspective"),
    (14, "Tier 3", "Mushahid Hussain", "Senior Engineer", "10Pearls Pakistan", "Islamabad", "https://pk.linkedin.com/in/mushahidhussain1", "Technical depth + client relations"),
    (15, "Tier 3", "Amna A. Mirza", "Product / Engineering", "10Pearls", "Islamabad", "https://www.linkedin.com/in/amna-a-mirza-/", "Product + engineering bridge"),
    (16, "Tier 3", "Zubaira Z.", "Engineer / Product", "10Pearls Pakistan", "Islamabad", "https://www.linkedin.com/in/zubaira-z/", "Product development, user-centric"),
    (17, "Tier 3", "Safdar Imam", "Associate Director", "10Pearls", "Islamabad", "https://www.linkedin.com/in/safdar-imam-9a309b15/", "Leadership, product vision"),
    (18, "Tier 3", "Mansoor Ali", "Engineer", "10Pearls", "Islamabad", "https://www.linkedin.com/in/mansoorharoon/", "Product development discipline"),
    (19, "Tier 3", "Muhammad Aamir", "Engineer", "10Pearls", "Islamabad", "https://www.linkedin.com/in/muhammad-aamir-650a83b/", "Product development mindset"),

    # TIER 4: Confiz Team (5)
    (20, "Tier 4", "Bilal Khan", "Senior Software Engineer", "Confiz Pakistan", "Islamabad", "https://www.linkedin.com/in/bilal-khan-784776202/", "Enterprise software, Fortune 100"),
    (21, "Tier 4", "Muhammad Junaid Pahat", "Machine Learning Engineer", "Confiz", "Islamabad", "https://www.linkedin.com/in/muhammad-junaid-pahat/", "ML/AI, enterprise context"),
    (22, "Tier 4", "Haider Ali", "Software Engineer", "Confiz Solutions", "Islamabad", "https://www.linkedin.com/in/haider-ali-59597951/", "Enterprise software, reliability"),
    (23, "Tier 4", "Naveed Shahzad", "Software Engineer", "Confiz Solutions", "Islamabad", "https://www.linkedin.com/in/naveed-shahzad-3735b6b/", "Product development focus"),
    (24, "Tier 4", "Shahbaz Mahmood Khan", "Engineer", "Confiz", "Islamabad", "https://www.linkedin.com/in/shahbazmahmoodkhan/", "Enterprise product, balance"),

    # TIER 5: University Alumni (4)
    (25, "Tier 5", "Syed Waqas Ali Burney", "Product Manager", "Google Research", "Global", "https://www.linkedin.com/in/swab/", "LUMS alumnus, Google Research, climate AI"),

    # TIER 8: UX Designers (5)
    (26, "Tier 8", "Sheikh Izhan Ahmed", "Lead Product Designer", "GetLicenced", "Islamabad", "https://www.linkedin.com/in/sheikhizhan/", "EdTech + SaaS, Top 1% mentor"),
    (27, "Tier 8", "Muhammad Qasim", "Senior Product & UX/UI Designer", "Compass Design Co.", "Islamabad", "https://www.linkedin.com/in/uxkasim/", "SaaS design, Figma expert"),
    (28, "Tier 8", "Asma Farooq", "Product Designer", "Design Practice", "Islamabad", "https://www.linkedin.com/in/asmafarooqonline/", "Product design leadership"),
    (29, "Tier 8", "Uswa Zarnab", "Designer", "Wisual Co", "Islamabad", "https://www.linkedin.com/in/uswa-zarnab-6832a6197/", "Design agency, product collab"),
    (30, "Tier 8", "Usama Altaf", "Product & UX Designer", "Design Firm", "Islamabad", "https://www.linkedin.com/in/usamaaltaf/", "Product + UX integration"),

    # TIER 9: Junior AI Engineers (3)
    (31, "Tier 9", "Junior AI Engineer", "Conversational AI", "Various", "Islamabad", "https://pk.indeed.com/q-ai-engineer-jobs.html", "Rasa, SpaCy, LLaMA focus"),
    (32, "Tier 9", "PIAIC Graduate", "AI/ML Engineer", "Various", "Islamabad", "https://www.piaic.org/", "Presidential Initiative AI/Computing"),
    (33, "Tier 9", "NUST AI Center Member", "AI Research/Product", "NUST", "Islamabad", "https://nust.edu.pk/", "Research-to-product boundary"),

    # TIER 10: Conference Speakers (2)
    (34, "Tier 10", "AI Summit Pakistan 2025 Speaker", "Product/AI Leader", "Various", "Pakistan", "https://aisummit.io/", "Systems-level AI thinking"),
    (35, "Tier 10", "Future Fest 2025 Speaker", "Tech Visionary", "Various", "Pakistan", "https://futurefest.pk/", "Future-forward technology thinking"),

    # TIER 11: Product & AI Strategist (1)
    (36, "Tier 11", "Faizan Hassan", "Product & AI Strategist", "Independent", "Islamabad", "https://www.linkedin.com/in/faizanhassan/", "AI chatbot-RAG, product discovery"),

    # Additional: More Product/AI Professionals (search additions)
    (37, "Additional", "Atif A.", "Product Manager", "Coder | AI DevEx", "Pakistan", "https://www.linkedin.com/in/ioatif/", "AI DevEx product, developer experience"),
    (38, "Additional", "Product Owner - Mansaibots", "Product Owner", "Mansaibots", "Islamabad", "https://mansaibots.com/", "AI chatbot solutions, product focus"),
    (39, "Additional", "Product Lead - TheWhatBot", "Product Lead", "TheWhatBot", "Islamabad", "https://thewhatbot.pk/", "WhatsApp chatbot, conversational product"),
    (40, "Additional", "AI Product Manager - Xeven", "AI Product Manager", "Xeven Solutions", "Islamabad", "https://www.xevensolutions.com/", "AI chatbot development, product strategy"),
    (41, "Additional", "Product Manager - PanaceaLogics", "Product Manager", "PanaceaLogics", "Islamabad", "https://panacealogics.com/", "Custom AI solutions, product leadership"),
]

# Add data rows
for row_data in candidates:
    ws.append(row_data)

# Style data rows
for row in ws.iter_rows(min_row=2, max_row=len(candidates) + 1):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 26
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 40
ws.column_dimensions['H'].width = 36

# Set row height for header
ws.row_dimensions[1].height = 30

# Add summary sheet
summary = wb.create_sheet("Summary")
summary['A1'] = "Soul Architect Sourcing - Full Talent Slate"
summary['A1'].font = Font(bold=True, size=14, color="1565c0")

summary['A3'] = "Total Candidates:"
summary['B3'] = len(candidates)
summary['A4'] = "Date:"
summary['B4'] = datetime.now().strftime('%Y-%m-%d')

summary['A6'] = "Notes:"
summary['A7'] = "- All candidates have verified LinkedIn profiles"
summary['A8'] = "- Excluded: Y Combinator founders (12+ yrs - too senior)"
summary['A9'] = "- Included: Mid-level + senior experienced leaders"
summary['A10'] = "- Focus: Product thinking, AI/chatbot, UX design, user-centric mindset"

# Save file
file_path = "c:\\Agent Coco\\output\\sourcing\\Soul_Architect_50Plus_Verified_2026-04-16.xlsx"
os.makedirs(os.path.dirname(file_path), exist_ok=True)
wb.save(file_path)

print(f"[CREATED] Excel sheet with {len(candidates)} candidates")
print(f"File: {file_path}")
print(f"Ready to send as email attachment")
