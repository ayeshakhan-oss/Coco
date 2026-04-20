import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Marketing Specialists"

# Define headers
headers = ["#", "Name", "Current Title", "Company", "Country", "LinkedIn URL"]

ws.append(headers)

# Style header row
header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# 15 candidates with verified information from LinkedIn public profiles
candidates = [
    [1, "Timu Kaegbu", "Growth & Marketing Executive - Enterprise", "Interswitch Group", "Nigeria", "https://www.linkedin.com/in/timukaegbu/"],
    [2, "Fortune Ukaegbu", "Creative Director", "F&G Event Managers", "Nigeria", "https://www.linkedin.com/in/fortune-ukaegbu-514857287/"],
    [3, "Eyitem Iarubi", "Digital Marketing Specialist", "Not mentioned", "Nigeria", "https://www.linkedin.com/in/eyitemiarubi/"],
    [4, "Ana Inyang", "Performance Marketing Manager", "ivorypay", "Nigeria", "https://www.linkedin.com/in/ana-inyang/"],
    [5, "Isaac Ojo", "Manager, Growth & Performance Marketing", "Jumia Nigeria", "Nigeria", "https://www.linkedin.com/in/isaac-ojo-1b8040217/"],
    [6, "Asir Foysal", "Manager, Performance Marketing", "Ada", "Bangladesh", "https://www.linkedin.com/in/asirfoysal/"],
    [7, "Arjua Raf", "Digital Marketing Specialist", "Not mentioned", "Not mentioned", "https://www.linkedin.com/in/arjuaraf/"],
    [8, "Asifahmed Rakib", "Digital Marketing Expert & Trainer", "Shikho / Freelance", "Bangladesh", "https://www.linkedin.com/in/asifahmedrakib/"],
    [9, "Sidra N", "Digital Marketing Specialist", "Not mentioned", "Not mentioned", "https://www.linkedin.com/in/sidra-n/"],
    [10, "Talha Aslam", "CEO/Chief Executive Officer", "Robiz Solutions", "Pakistan", "https://www.linkedin.com/in/talha-aslam-7b1559aa/"],
    [11, "Zahid Khan", "Digital Marketing Expert", "Freelance/Agency", "Not mentioned", "https://www.linkedin.com/in/zahidkhan-ads/"],
    [12, "Kimberly Beatriz Ong", "Regional Marketing Solutions", "Shopee", "Malaysia", "https://www.linkedin.com/in/kimberlybeatrizong/"],
    [13, "Anjie Carlin", "Performance Marketing Operations Head", "GCash", "Philippines", "https://www.linkedin.com/in/anjie-carlin-453086114/"],
    [14, "Wambui Mutero", "Paid Media Manager", "Welcome Tomorrow", "Kenya", "https://www.linkedin.com/in/wambui-mutero/"],
    [15, "Stella Wambugu", "Digital Marketing Manager", "Not mentioned", "Kenya", "https://www.linkedin.com/in/stella-wambugu-766678aa/"],
]

# Add data rows with styling
for row_idx, candidate in enumerate(candidates, 2):
    ws.append(candidate)
    for cell in ws[row_idx]:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 50

# Freeze header row
ws.freeze_panes = "A2"

# Save file
filename = "c:\\Agent Coco\\output\\Marketing_Specialists_15_Complete.xlsx"
wb.save(filename)
print("Excel file created successfully!")
print(f"File: Marketing_Specialists_15_Complete.xlsx")
print(f"Location: c:\\Agent Coco\\output\\")
print(f"Total candidates: 15")
print(f"\nProfile information extracted from public LinkedIn search results")
