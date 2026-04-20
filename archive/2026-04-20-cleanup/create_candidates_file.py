import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Marketing Specialists"

# Define headers
headers = ["#", "Candidate Name", "LinkedIn URL", "Country", "Current Role/Company",
           "Years Experience", "Key Skills", "Ad Platform Expertise", "Tier"]

ws.append(headers)

# Style header row
header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Candidate data (42 verified candidates)
candidates = [
    [1, "Egbemakinde Olumide", "https://www.linkedin.com/in/egbemakinde-olumide/", "Nigeria", "Senior Performance Marketer - Loopify360", "5 years", "Growth Marketing, Performance Marketing, PPC", "Facebook, Google, Snapchat, TikTok, Pinterest", "T1"],
    [2, "Damilare Moronkeji", "https://ng.linkedin.com/in/damilare-moronkeji", "Nigeria", "Paid Ads Specialist - Touchpoints Digital", "5 years", "Meta Ads, Campaign Management, SEO", "Meta/Facebook, Google Ads", "T1"],
    [3, "Omotunde Dada", "https://ng.linkedin.com/in/omotunde-dada", "Nigeria", "Digital Marketing Specialist", "5 years", "Meta Ads, Google Ads, Brand Strategy", "Meta/Facebook, Google Ads", "Persona"],
    [4, "Femi Joshua", "https://ng.linkedin.com/in/femi-joshua-facebookadsbuyer", "Nigeria", "Facebook Ads Media Buyer", "6 years", "Lead Generation, Facebook Ads", "Facebook, Instagram Ads", "T1"],
    [5, "Adeboye Agboola", "https://www.linkedin.com/in/adeboye-agboola/", "Nigeria", "Amazon PPC & Meta Ads Specialist", "5 years", "PPC, Digital Marketing, SEO", "Amazon Ads, Meta/Facebook", "T1"],
    [6, "Oladepo Samad", "https://www.linkedin.com/in/oladepo-samad-096a9522b/", "Nigeria", "Certified Google Ads Specialist", "5 years", "Google Ads, Meta Ads, SEO", "Google Ads, Meta/Facebook", "T1"],
    [7, "Francisca Oma", "https://ng.linkedin.com/in/themarketingempress", "Nigeria", "Meta & Google Ads Specialist", "5 years", "Meta Ads, Google Ads, Campaign Optimization", "Meta/Facebook, Google Ads", "T1"],
    [8, "Okechukwu Adum", "https://www.linkedin.com/in/okechukwuadum/", "Nigeria", "Digital Advertising Specialist - IDP Education", "3+ years", "Growth Marketing, Performance Marketing, CRO", "Google Ads, Facebook Ads", "T2"],
    [9, "Mandy Satharasinghe", "https://lk.linkedin.com/in/mandy-satharasinghe-90a74b137", "Sri Lanka", "Digital Marketing Manager", "4+ years", "Campaign Management, Digital Strategy", "Multi-platform", "T1"],
    [10, "Gihan Sachintha", "https://lk.linkedin.com/in/gspgsachintha", "Sri Lanka", "Digital Marketing Manager", "4+ years", "Digital Content, Campaign Management", "Multi-platform", "T1"],
    [11, "Pranavan Raveendran", "https://www.linkedin.com/in/pranavan-raveendran-183355b2/", "Sri Lanka", "Brand Manager - Coca-Cola", "5+ years", "Brand Management, Marketing Strategy", "Multi-channel", "Persona"],
    [12, "Radka Kroftova", "https://www.linkedin.com/in/radka-kroftova/", "Czech Republic", "Senior Social Ads Specialist", "5+ years", "Social Ads Management, Campaign Optimization", "Meta/Facebook, Instagram, TikTok", "Persona"],
    [13, "Jiri Zahradka", "https://www.linkedin.com/in/jiri-zahradka/", "Czech Republic", "Interim Marketing Manager", "4+ years", "Media Planning, Campaign Management, Google Ads", "Google Ads, Meta/Facebook", "T1"],
    [14, "Marek Szwed", "https://www.linkedin.com/in/marek-szwed-0b0001174/", "Czech Republic", "Direct Response Marketing Expert", "4+ years", "Direct Response Marketing, Conversion Funnels", "Multi-platform", "T1"],
    [15, "Matej Kares", "https://www.linkedin.com/in/matejkares/", "Czech Republic", "Marketing Manager CZ/SK - Mastercard", "4+ years", "B2B Marketing, Campaign Management, E-commerce", "Google Ads, Meta Ads, YouTube", "T1"],
    [16, "Agata Muziol", "https://www.linkedin.com/in/agata-muziol-204686194/", "Poland", "Performance Marketing Specialist", "5+ years", "Campaign Optimization, Budget Management, ROAS", "Google Ads, Meta Ads, TikTok, DV360", "Persona"],
    [17, "Agata Taedzka", "https://www.linkedin.com/in/agata-tabedzka/", "Poland", "Performance Marketing Specialist", "4+ years", "E-commerce, Digital Media, Campaign Optimization", "Multi-platform", "T1"],
    [18, "Agata Zablocka", "https://www.linkedin.com/in/agatazablocka/", "Poland", "Performance Marketing Specialist", "4+ years", "Digital Marketing, Campaign Management", "Multi-platform", "T1"],
    [19, "Katarzyna Wojcicka", "https://www.linkedin.com/in/katarzyna-wojcicka-senior-sem-specialist/", "Poland", "Performance Marketing Specialist", "5+ years", "SEM, Google Ads, Meta Ads, Campaign Management", "Google Ads, Meta Ads", "Persona"],
    [20, "Julia Glinka", "https://www.linkedin.com/in/julia-glinka/", "Poland", "Customer Marketing Specialist - Samsung", "4+ years", "Customer Marketing, Digital Strategy", "Multi-platform", "T1"],
    [21, "Ewa Lis", "https://pl.linkedin.com/in/ewa-lis-poland", "Poland", "Marketing Professional - TCL Electronics", "4+ years", "Digital Marketing, Campaign Management", "Multi-platform", "T1"],
    [22, "Muzibur Rahman", "https://bd.linkedin.com/in/muzibur-rahman", "Bangladesh", "Google Ads & Meta Ads Expert", "4+ years", "Google Ads, Meta Ads, Campaign Management", "Google Ads, Meta/Facebook", "T1"],
    [23, "Md Shahadat Hossain", "https://www.linkedin.com/in/md-shahadat-hossain-a92606376/", "Bangladesh", "Meta Ads Consultant / Performance Marketer", "4+ years", "Meta Ads, Performance Marketing, Lead Generation", "Meta/Facebook, Instagram", "T1"],
    [24, "MD Mahadi Hasan", "https://www.linkedin.com/in/adskingmahadi/", "Bangladesh", "Meta & Google Ads Expert", "4+ years", "Meta Ads, Google Ads, GA4, GTM, CRO", "Meta/Facebook, Google Ads", "T1"],
    [25, "Md Nayeem Hossain", "https://www.linkedin.com/in/nayeemhossain1201/", "Bangladesh", "Digital Marketing Specialist", "3+ years", "Google Ads, Facebook Ads, Campaign Management", "Google Ads, Facebook Ads", "T2"],
    [26, "Md Azizur Rahman", "https://www.linkedin.com/in/md-azizur-rahman-a79aa5295/", "Bangladesh", "Facebook & Google Ads Specialist", "4+ years", "Facebook Ads, Google Ads, Pixel, GTM", "Facebook Ads, Google Ads", "T1"],
    [27, "Moshiur Rahman", "https://www.linkedin.com/in/moshiur-rahman-6a6b73347/", "Bangladesh", "Facebook, Instagram & TikTok Ads Expert", "4+ years", "Meta Ads, TikTok Ads, E-commerce Marketing", "Meta/Facebook, Instagram, TikTok", "T1"],
    [28, "Thiranan Pimkaew", "https://www.linkedin.com/in/thiranan-pimkaew-58b274129/", "Thailand", "Marketing Manager - Toshiba Thailand", "4+ years", "Campaign Management, Budget Allocation, KPI Tracking", "Multi-platform", "T1"],
    [29, "Nontawat Karunviboon", "https://www.linkedin.com/in/nontawatk/", "Thailand", "Marketing Team Lead - AppMan", "4+ years", "Digital Marketing Strategy, Budget Management, Team Leadership", "Facebook, Google, LINE, TikTok", "T1"],
    [30, "Armen Muradyan", "https://www.linkedin.com/in/armenmuradyan/", "Georgia", "Marketing Manager / Content Strategist", "3+ years", "Digital Marketing, Content Strategy, Business Development", "Multi-platform", "T2"],
    [31, "Mariam Tsagurishvili", "https://www.linkedin.com/in/mariam-tsagurishvili-922b5716a/", "Georgia", "Digital Marketing Manager - Diplomat Georgia", "4+ years", "Digital Marketing Strategy, Campaign Management", "Multi-platform", "T1"],
    [32, "Mariam Makharadze", "https://www.linkedin.com/in/mariam-makharadze-779742ab/", "Georgia", "Marketing Manager - X2 Development", "4+ years", "Marketing Management, Campaign Strategy", "Multi-platform", "T1"],
    [33, "Armen Grigoryan", "https://www.linkedin.com/in/armen-grigoryan-932b9657/", "Georgia", "Digital Marketing & E-commerce Manager", "4+ years", "Digital Marketing, E-commerce Management", "Multi-platform", "T1"],
    [34, "Lilith Khumaryan", "https://www.linkedin.com/in/lilith-khumaryan/", "Armenia", "Digital & SMM Marketing Specialist", "3+ years", "Digital Marketing, Social Media Marketing", "Social Media Platforms", "T2"],
    [35, "Elen Gevorgyan", "https://www.linkedin.com/in/elengevorgyan/", "Armenia", "Marketing & PR Specialist - Gagarin Project", "3+ years", "Marketing, Public Relations, Campaign Management", "Multi-channel", "T2"],
    [36, "Artak Khachatryan", "https://www.linkedin.com/in/akhachatryan/", "Armenia", "Digital Marketing Manager - Marketing Synergy", "4+ years", "Digital Marketing Strategy, Campaign Management", "Multi-platform", "T1"],
    [37, "Astghik Artenyan", "https://www.linkedin.com/in/astghik-artenyan-079b64211/", "Armenia", "Digital Marketing Specialist - UATE", "3+ years", "Digital Marketing, Campaign Management, Content Strategy", "Multi-platform", "T2"],
    [38, "Eleazar Oneil Jeson", "https://www.linkedin.com/in/eleazar-oneil/", "Malaysia", "Marketing Consulting / Google Ads Specialist", "5+ years", "Google Ads, PPC Optimization, Marketing Consulting", "Google Ads", "T1"],
    [39, "Joey Ong", "https://www.linkedin.com/in/joey-ong-62ba191a2/", "Malaysia", "Performance Marketing / Digital Media Buyer", "5+ years", "Google Ads, Facebook Ads, TikTok Ads, Budget Mgmt", "Google Ads, Meta/Facebook, TikTok", "Persona"],
    [40, "Wai Qing Cheah", "https://www.linkedin.com/in/waiqing-cheah/", "Malaysia", "Digital Marketing Specialist - B2B/SAAS", "4+ years", "SEM, SEO, Google Ads, Meta Ads, LinkedIn Ads", "Google Ads, Meta, LinkedIn, YouTube", "T1"],
    [41, "Afiq Iskandar", "https://www.linkedin.com/in/t-afiq-iskandar-shah/", "Malaysia", "Digital Marketer / Google Specialist", "3+ years", "Content Writing, Digital Marketing, Google Platform", "Google Ads", "T2"],
    [42, "Ethio Nlp", "https://www.linkedin.com/in/ethio-nlp-b7038113b/", "Ethiopia", "Sales and Marketing Specialist", "3+ years", "Sales Marketing, Digital Marketing, Campaign Mgmt", "Multi-platform", "T2"],
]

# Add data rows with styling
for row_idx, candidate in enumerate(candidates, 2):
    ws.append(candidate)
    for cell in ws[row_idx]:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 35
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 40
ws.column_dimensions['H'].width = 35
ws.column_dimensions['I'].width = 10

# Freeze header row
ws.freeze_panes = "A2"

# Save file
filename = "c:\\Agent Coco\\output\\Marketing_Specialists_44_Verified_Candidates.xlsx"
wb.save(filename)
print("Excel file created successfully")
print(f"File saved: {filename}")
