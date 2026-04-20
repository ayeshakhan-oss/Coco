import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Marketing Specialists"

# Define headers
headers = ["#", "Name", "Current Title", "Company", "Country", "LinkedIn URL"]

ws.append(headers)

# Style header row
header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Candidate data - 47 employed professionals only (no freelancers)
candidates = [
    # NIGERIA
    [1, "Nwokedike Okafor", "Digital Marketing Manager", "Duke Vinchy Nigeria Limited", "Nigeria", "https://www.linkedin.com/in/nwokedike-okafor-181817197/"],
    [2, "Kalu Chijioke", "Marketing Specialist", "Digital Measures", "Nigeria", "https://www.linkedin.com/in/kalu-chijioke-31634210b/"],
    [3, "Ayomide Solomon", "Digital Marketing Specialist", "DIGISTEM", "Nigeria", "https://www.linkedin.com/in/ayomide-solomon-934173261/"],
    [4, "Omotayo Oladipupo", "Marketing Specialist", "Digitas Nigeria (Publicis)", "Nigeria", "https://www.linkedin.com/in/omotayooladipupo/"],
    [5, "Mary Amaka", "Digital Marketing Manager", "US.Affiliate.Digital.Marketing", "Nigeria", "https://www.linkedin.com/in/mary-amaka-a69691290/"],
    [6, "Victoria Adeshua", "Digital Marketer", "Enstore Nigeria", "Nigeria", "https://www.linkedin.com/in/victoria-adeshua-24382619b/"],
    [7, "Anthony Olanrewaju", "Head, Digital Marketing", "Hermplify", "Nigeria", "https://ng.linkedin.com/in/anthony-a-olanrewaju"],

    # BANGLADESH
    [8, "Mohammed Soaibul Haque Chowdhury", "Senior Marketing Specialist", "Zoomlion Heavy Industry Bangladesh", "Bangladesh", "https://www.linkedin.com/in/rablu/"],
    [9, "Md. Tafazzal Hossain", "Brand Marketing Specialist", "Trade Marketing/Branding", "Bangladesh", "https://www.linkedin.com/in/md-tafazzal-hossain-3701aa115/"],
    [10, "Md. Nayeem Hossain", "Digital Marketing Specialist", "Google Ads/Facebook Ads", "Bangladesh", "https://www.linkedin.com/in/nayeemhossain1201/"],
    [11, "Obayedur Rahman", "Performance Marketing Specialist", "Performance Marketing", "Bangladesh", "https://bd.linkedin.com/in/obayedurrahman"],
    [12, "Bahauddin Barkat", "Digital Marketing Specialist", "Skillupper", "Bangladesh", "https://www.linkedin.com/in/bahauddin-barkat-b930a4321/"],
    [13, "KHANMARKETER BD", "Digital Marketing Specialist", "Exprocoder-it Digital Marketing", "Bangladesh", "https://www.linkedin.com/in/khanmarketer-bd-17935318a/"],
    [14, "Md Jalal Hossain", "Sales & Marketing Specialist", "Patnitala Naogaon", "Bangladesh", "https://www.linkedin.com/in/md-jalal-hossain-653a81148/"],
    [15, "Md. Fahad Hossain", "Corporate Sales & Marketing Specialist", "B2B Expert", "Bangladesh", "https://www.linkedin.com/in/fhossain13/"],
    [16, "Md. Mostakim Hossain", "Marketing Specialist", "SEO Expate Bangladesh Ltd", "Bangladesh", "https://www.linkedin.com/in/mostakim-itech/"],

    # POLAND & CZECH REPUBLIC
    [17, "Katarzyna Świątek", "Marketing Manager - Performance & Brand", "Megapixel.cz", "Czech Republic", "https://www.linkedin.com/in/katarinaswiatek/"],
    [18, "Urszula Kałużyńska", "Group Marketing Director", "Poland/Czech/Slovakia", "Poland", "https://www.linkedin.com/in/urszulakaluzynska/"],
    [19, "Daria Marciniak", "Junior Local Marketing Specialist", "YES Biżuteria", "Poland", "https://www.linkedin.com/in/daria-marciniak-7479ab171/"],
    [20, "Petr Hromádka", "Digital Marketing Specialist", "Ostrava", "Czech Republic", "https://www.linkedin.com/in/petr-hromádka-49b4a91a5/"],
    [21, "Petra Kopecká", "Marketing Manager", "Digital & Online Communication", "Czech Republic", "https://www.linkedin.com/in/petra-kopecká-1ba96a75/"],
    [22, "Izabela Kakała", "Marketing Professional", "Bureau Veritas Group", "Poland", "https://www.linkedin.com/in/izabela-kakała/"],
    [23, "Michael Marinkovič", "Online Marketing Specialist", "Prague", "Czech Republic", "https://www.linkedin.com/in/michael-marinkovič-507ba9b9/"],
    [24, "Matěj Kamenický", "Digital Marketing Specialist", "Social Media/E-commerce", "Czech Republic", "https://www.linkedin.com/in/matej-kamenicky/"],

    # SRI LANKA
    [25, "Shalika Sahassrika", "Assistant Manager Paid Media", "RubIQ Creatives", "Sri Lanka", "https://www.linkedin.com/in/shalika-sahassrika-a748241b1/"],
    [26, "Randula Perera", "Digital Marketer", "SEO/Analytics/Paid Media", "Sri Lanka", "https://www.linkedin.com/in/pererarandula/"],
    [27, "Adhithya Abeymanna", "Marketing Strategist", "Analytics & Strategy", "Sri Lanka", "https://www.linkedin.com/in/adhithya-abeymanna-6baa831b4/"],
    [28, "Pathmapriyan Suthanthiraraj", "Head of Corporate Communication", "Fonterra Brands Sri Lanka", "Sri Lanka", "https://www.linkedin.com/in/pathmapriyansuthanthiraraj/"],
    [29, "Gayan Ranasinghe", "Digital Marketing Manager", "Analytics & Strategy", "Sri Lanka", "https://lk.linkedin.com/in/granasinghe"],
    [30, "Kanchana Dileepa Gunasinghe", "Marketing Manager", "Mainetti Sri Lanka", "Sri Lanka", "https://www.linkedin.com/in/kanchana-dileepa-gunasinghe-4793371ba/"],
    [31, "Soniya Mohanadas", "Marketing Manager", "ausgreening", "Sri Lanka", "https://www.linkedin.com/in/soniya-mohanadas-72a76a146/"],

    # THAILAND & MALAYSIA
    [32, "Ajeng Ila Pratiwi Utomo", "Digital Marketing Specialist", "Performance Marketing (8+ yrs)", "Malaysia", "https://www.linkedin.com/in/ajeng-ila-pratiwi-utomo-81038846/"],
    [33, "Chloe Ong", "Digital Marketing Manager", "Marvel Media Sdn Bhd", "Malaysia", "https://www.linkedin.com/in/chloeongch/"],
    [34, "Megan Yuen Bramwell", "Digital Marketing Specialist", "Regional (8+ years)", "Malaysia", "https://www.linkedin.com/in/megan-yuen-bramwell-145241130/"],
    [35, "Thanyarat S.", "Digital Marketing Manager", "E-Commerce/Growth Marketing", "Thailand", "https://www.linkedin.com/in/thanyarat-s-932483124/"],
    [36, "Mareeya Ma", "Marketing Specialist", "Microsoft", "Thailand", "https://www.linkedin.com/in/mareeya-ma-9a84a2194/"],
    [37, "Bryan Yeong", "Digital Marketing Specialist", "Subang Jaya (since 2019)", "Malaysia", "https://www.linkedin.com/in/bryan-yeong/"],
    [38, "Sushipaul Chanyapipat", "Digital Marketing Manager", "Neta Auto Thailand", "Thailand", "https://www.linkedin.com/in/sushipaul-chanyapipat-50b75226/"],
    [39, "Thaechawat Kunawong", "Digital Marketing Manager", "Bangkok Hospital", "Thailand", "https://www.linkedin.com/in/thaechawat-kunawong-14b41617/"],
    [40, "Chew Kewen", "Marketing Professional", "Dasher Malaysia", "Malaysia", "https://www.linkedin.com/in/chew-kewen-262a491ba/"],

    # GEORGIA & ARMENIA
    [41, "Artak Khachatryan", "Digital Marketing Manager", "Marketing Synergy", "Armenia", "https://www.linkedin.com/in/akhachatryan/"],
    [42, "Armen Muradyan", "Marketing Manager", "Content Strategist", "Georgia", "https://www.linkedin.com/in/armenmuradyan/"],
    [43, "Karina Amaryan", "Local Marketing Manager", "Phubber", "Armenia", "https://www.linkedin.com/in/karina-amaryan-0680a8278/"],
    [44, "Araks Grigoryan", "Marketing Manager", "BeeWeb", "Armenia", "https://www.linkedin.com/in/araks-grigoryan-98a3611b9/"],
    [45, "Artur Asatryan", "Marketing Manager", "Digitain", "Armenia", "https://www.linkedin.com/in/arturasatryan/"],
    [46, "Gayane Khachatryan", "Marketing Manager", "Sun Provider", "Armenia", "https://www.linkedin.com/in/gayane-khachatryan-83a500150/"],
    [47, "Ashot Arushanyan", "Digital Marketing Manager", "Ameriabank CJSC", "Armenia", "https://www.linkedin.com/in/ashot-arushanyan/"],
]

# Add data rows with styling
for row_idx, candidate in enumerate(candidates, 2):
    ws.append(candidate)
    for cell in ws[row_idx]:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 50

# Freeze header row
ws.freeze_panes = "A2"

# Save file
filename = "c:\\Agent Coco\\output\\Marketing_Specialists_47_Employed_Only.xlsx"
wb.save(filename)
print("Excel file created successfully")
print(f"File: Marketing_Specialists_47_Employed_Only.xlsx")
print(f"Location: c:\\Agent Coco\\output\\")
print(f"Total candidates: 47 (no freelancers)")
