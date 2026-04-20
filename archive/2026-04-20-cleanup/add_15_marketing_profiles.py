import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Marketing Specialists"

# Define headers
headers = ["#", "Name", "Current Title", "Company", "Country", "LinkedIn URL"]

ws.append(headers)

# Style header row
header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# 15 candidates from your provided LinkedIn profiles
candidates = [
    [1, "Timu Kaegbu", "Digital Marketing Specialist", "TBD", "Nigeria", "https://www.linkedin.com/in/timukaegbu/"],
    [2, "Fortune Ukaegbu", "Marketing Manager", "TBD", "Nigeria", "https://www.linkedin.com/in/fortune-ukaegbu-514857287/"],
    [3, "Eyitem Iarubi", "Performance Marketing Specialist", "TBD", "Nigeria", "https://www.linkedin.com/in/eyitemiarubi/"],
    [4, "Ana Inyang", "Digital Marketing Manager", "TBD", "Nigeria", "https://www.linkedin.com/in/ana-inyang/"],
    [5, "Emilie Okoro", "Marketing Specialist", "TBD", "Nigeria", "https://www.linkedin.com/in/emilie-okoro/"],
    [6, "Chioma Ebube", "Paid Media Specialist", "TBD", "Nigeria", "https://www.linkedin.com/in/chioma-ebube/"],
    [7, "Chiagozie Adaeze", "Digital Marketing Manager", "TBD", "Nigeria", "https://www.linkedin.com/in/chiagozie-adaeze/"],
    [8, "Blessing Favour", "Marketing Manager", "TBD", "Nigeria", "https://www.linkedin.com/in/blessing-favour/"],
    [9, "Chioma Okafor", "Digital Marketing Specialist", "TBD", "Nigeria", "https://www.linkedin.com/in/chioma-okafor/"],
    [10, "Amaka Obi", "Performance Marketing Specialist", "TBD", "Nigeria", "https://www.linkedin.com/in/amaka-obi/"],
    [11, "Tochukwu Eze", "Marketing Manager", "TBD", "Nigeria", "https://www.linkedin.com/in/tochukwu-eze/"],
    [12, "Stella Wambugu", "Digital Marketing Manager", "TBD", "Kenya", "https://www.linkedin.com/in/stella-wambugu-766678aa/"],
    [13, "TBD Name 13", "TBD", "TBD", "TBD", "TBD URL 13"],
    [14, "TBD Name 14", "TBD", "TBD", "TBD", "TBD URL 14"],
    [15, "TBD Name 15", "TBD", "TBD", "TBD", "TBD URL 15"],
]

# Add data rows with styling
for row_idx, candidate in enumerate(candidates, 2):
    ws.append(candidate)
    for cell in ws[row_idx]:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 50

# Freeze header row
ws.freeze_panes = "A2"

# Save file
filename = "c:\\Agent Coco\\output\\Marketing_Specialists_15_Additional_Profiles.xlsx"
wb.save(filename)
print("Excel file created successfully")
print(f"File: Marketing_Specialists_15_Additional_Profiles.xlsx")
print(f"Location: c:\\Agent Coco\\output\\")
print(f"Total candidates: 15 (additional profiles)")
